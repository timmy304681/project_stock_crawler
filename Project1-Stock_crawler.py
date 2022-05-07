#Python網頁爬蟲與MySQL資料庫的整合方式

from bs4 import BeautifulSoup
import requests
import lxml
import pymysql
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font

class Stock:
    #建構式
    def __init__(self, *stock_numbers):
        self.stock_numbers = stock_numbers
        print(self.stock_numbers)

    #爬取
    def crawler(self):

        result = list()  # 最終結果

        for stock_number in self.stock_numbers:
            response = requests.get("https://tw.stock.yahoo.com/quote/"+stock_number)
            soup = BeautifulSoup(response.text, "lxml")

            #利用「公司名稱」的<h1>標籤、「資料時間」的<span>標籤，以及各自的樣式類別(class)，即可進行元素的定位，
            #再透過BeautifulSoup套件的getText()方法就可以取得其中的資料
            stock_name = soup.find('h1', {'class': 'C($c-link-text) Fw(b) Fz(24px) Mend(8px)'}).getText()

            stock_date = soup.find('span', {'class': 'C(#6e7780) Fz(14px) As(c)'}).getText().replace('資料時間：', '')
            market_date = stock_date[0:10]  #日期
            market_time = stock_date[11:16]  #時間

            #利用<ul>標籤及它的樣式類別(class)定位,取得「當日行情資料」清單
            ul = soup.find('ul', {'class': 'D(f) Fld(c) Flw(w) H(192px) Mx(-16px)'})
            items = ul.find_all('li', {'class': 'price-detail-item H(32px) Mx(16px) D(f) Jc(sb) Ai(c) Bxz(bb) Px(0px) Py(4px) Bdbs(s) Bdbc($bd-primary-divider) Bdbw(1px)'})
            # 讀取每一個項目(li)下的第2個<span>標籤資料值，並且存放在元組(Tuple)中
            data = tuple(item.find_all('span')[1].getText() for item in items)  #數據
            tilte= tuple(item.find_all('span')[0].getText() for item in items)  #數據名稱

            #整理資料
            result.append((market_date, stock_name, market_time) + data)
        return result

    #將檔案寫進mysql
    #mysql db資料
    def save2sql(self,stocks):
        db_settings = {
            "host": "localhost",
            "port": 3306,
            "user": "root",
            "password": password,
            "db": "stock_db",
            "charset": "utf8"}
        try:
            # 使用pymysql的cursor使資料傳入sql
            conn = pymysql.connect(**db_settings)

            with conn.cursor() as cursor:
                #column name要與mysql一致
                sql = """INSERT INTO market(
                    market_date,
                    stock_name,
                    market_time,
                    final_price,
                    opening_price,
                    highest_price,
                    lowest_price,
                    average_price,
                    transaction_value,
                    yesterday_price,
                    quote_change,
                    ups_and_downs,
                    total_volume,
                    yesterday_volume,
                    amplitude)
                    VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

                #將每一支股票資料傳入MySQL中
                for stock in stocks:
                    cursor.execute(sql, stock)
                conn.commit()

        except Exception as ex:
            print("Exception:", ex)

    #將資料存入google sheet
    def gsheet(self, stocks):
        scopes = ["https://spreadsheets.google.com/feeds"]

        credentials = ServiceAccountCredentials.from_json_keyfile_name(
                    "credits_googlesheet.json", scopes)   #credits_googlesheet.json要先建起google sheet api，並取得憑證
        #使用憑證取得google sheet api授權
        client = gspread.authorize(credentials)

        sheet = client.open_by_key(googlesheetkey).sheet1 #sheet1為第一個工作表

        for stock in stocks:
            sheet.append_row(stock)


    #輸出到excel
    def export2excel(self, stocks):
        wb = openpyxl.Workbook() #建立EXCEL
        sheet = wb.create_sheet("Yahoo股市", 0) #sheet 名稱

        #爬取標題，選隨意一股
        response = requests.get("https://tw.stock.yahoo.com/q/q?s=2451")
        soup = BeautifulSoup(response.text, "lxml")
        ul = soup.find('ul', {'class': 'D(f) Fld(c) Flw(w) H(192px) Mx(-16px)'})
        items = ul.find_all('li', {'class': 'price-detail-item H(32px) Mx(16px) D(f) Jc(sb) Ai(c) Bxz(bb) Px(0px) Py(4px) Bdbs(s) Bdbc($bd-primary-divider) Bdbw(1px)'})
        # 讀取每一個項目(li)下的第2個<span>標籤資料值，並且存放在元組(Tuple)中
        colnames= tuple(item.find_all('span')[0].getText() for item in items)  #數據名稱
        titles = ("資料日期","股票","時間") + colnames
        sheet.append(titles)

        #寫入資料
        for stock in stocks:
            sheet.append(stock)

            if "△" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='FF0000')  #儲存格字體顯示紅色
            elif "▽" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='00A600')  #儲存格字體顯示綠色

        #資料存檔
        wb.save("yahoostock.xlsx")




##### 執行
#輸入想關注的目標股票
stock = Stock("2330","2303","2603","2609")  #建立Stock物件

##Set password & keys
password="YOUR PASSWORD"
googlesheetkey="YOUR GOOGLE SHEET KEY"


##Start
print(stock.crawler())
#stock.save2sql(stock.crawler())    #將資料存入mysql
stock.gsheet(stock.crawler())       #將資料存入google sheet
stock.export2excel(stock.crawler()) #將資料寫入 excel